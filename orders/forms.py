import datetime

from django.forms import ModelForm
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from django.utils.text import slugify

from .models import OrderLog, Order, Operation
from django import forms


class OrderLogForm(ModelForm):
    ITERATION_OPERATIONS = [
        "gruboe-volochenie",
        "liniya-70",
        "lentoobmotka",
    ]
    BUHTOVKA = [
        "buhtovka",
    ]

    def __init__(self, *args, **data):
        super(OrderLogForm, self).__init__(*args, **data)
        if len(self.initial) and data:
            self.fields["order"].queryset = Order.objects.filter(
                id=data["initial"]["order"].id
            )
        self.fields["iteration"].widget = forms.HiddenInput()
        self.form_for_buhtovka(data) if len(data) else None
        self.get_hidden_fields(data) if len(data) else None

    id_order_log = forms.IntegerField(widget=forms.HiddenInput)

    class Meta:
        model = OrderLog
        fields = [
            "order",
            "operation",
            "operator",
            "color_cores",
            "prev_number_container",
            "container",
            "number_container",
            "total_in_meters",
            "comment",
            "iteration",
        ]

    def get_operation(self, data):
        if len(self.initial):
            operation = data["initial"]["operation"]
        else:
            operation_id = [
                val
                for field in data.values()
                for key, val in field.items()
                if key == "operation"
            ]
            operation = get_object_or_404(Operation, id=operation_id[0])
        return operation

    def form_for_buhtovka(self, data):
        operation = self.get_operation(data)
        if operation.slug in self.BUHTOVKA:
            self.fields["color_cores"].initial = "б/ц"
            self.fields["container"].initial = "бух"
            self.fields["number_container"].label = "Кол-во бухт"
            self.fields["total_in_meters"].label = "Метраж бухты"

    def get_hidden_fields(self, data):
        operation = self.get_operation(data)
        hidden_field_3 = ["gruboe-volochenie", "bolshaya-skrutka", "lentoobmotka"]
        hidden_field_2 = ["liniya-90", "buhtovka"]
        hidden_field_1 = ["liniya-70"]
        if operation.slug in hidden_field_3:
            self.fields["color_cores"].widget = forms.HiddenInput()
            self.fields["prev_number_container"].widget = forms.HiddenInput()
            self.fields["container"].widget = forms.HiddenInput()
        elif operation.slug in hidden_field_2:
            self.fields["color_cores"].widget = forms.HiddenInput()

        elif operation.slug in hidden_field_1:
            self.fields["prev_number_container"].widget = forms.HiddenInput()
            self.fields["container"].widget = forms.HiddenInput()

    def clean_total_in_meters(self):
        cd = self.cleaned_data
        order = cd["order"]
        operation = cd["operation"]
        total_meters = cd["total_in_meters"]
        order_log_meter = OrderLog.objects.filter(
            order=order, operation=operation
        ).values_list("total_in_meters", flat=True)
        order_total_meter = sum(order_log_meter) + cd["total_in_meters"]
        if operation.slug in self.ITERATION_OPERATIONS:
            # Проверка на превышение длинны
            if int(order.footage) + 200 < total_meters:
                raise forms.ValidationError(
                    f"Метраж указан больше чем в заказе. Длинна заказа {order.footage} м."
                )
            # Проверка, если длинна меньше требуемой
            elif int(order.footage) - 200 > total_meters:
                raise forms.ValidationError(
                    f"Указанная длинна меньше длинны заказа. Длинна заказа {order.footage} м."
                )
        elif operation.slug in self.BUHTOVKA:
            len_bight = cd["number_container"]
            order_log_meter = OrderLog.objects.filter(
                order=order, operation=operation
            ).values_list("total_in_meters", "number_container")
            if order_log_meter:
                order_total_meter = sum(map(lambda x: x[0] * x[1], order_log_meter))
            if total_meters > 200:
                raise forms.ValidationError(f"Длинна бухты не должна превышать 200 м.")
            elif int(order.footage) + 200 < total_meters * len_bight:
                raise forms.ValidationError(
                    f"Превышение общей длинны заказа. Заказ {order.footage}м. Выход {len_bight * total_meters}м. Бухта {len_bight}м. Длинна бухт {total_meters}м."
                )
            elif (
                (int(order.footage) + 200)
                - (order_total_meter + (len_bight * total_meters))
            ) < 0:
                raise forms.ValidationError(
                    f"Метраж указан больше чем в заказе. Длинна заказа {order.footage} м. Остаток {order.footage - order_total_meter}"
                )
        else:
            # Проверка обшей длинны заказа при делении заказа на барабаны
            if (int(order.footage) + 200) - order_total_meter < 0:
                raise forms.ValidationError(
                    f"Метраж указан больше чем в заказе. Длинна заказа {order.footage} м. Остаток {order.footage - (sum(order_log_meter))}"
                )
        return cd["total_in_meters"]

    def clean_number_container(self):
        cd = self.cleaned_data
        number_container = cd["number_container"]
        order = cd["order"]
        operation = cd["operation"]
        # Проверка номера катушки
        if number_container > 250 and operation.slug not in self.BUHTOVKA:
            raise forms.ValidationError("Проверь номер катушки")
        num_containers_in_operation = OrderLog.objects.filter(
            order=order, operation=operation
        ).values_list("number_container", flat=True)
        # Проверка на дублирование номера катушки
        if (
            number_container in num_containers_in_operation
            and operation.slug not in self.BUHTOVKA
        ):
            raise forms.ValidationError(
                f"Катушка №{number_container} уже присутствует в заказе"
            )
        return cd["number_container"]


class UploadDataOrdersForm(forms.Form):
    header = [
        "batch_number",
        "plastic",
        "design",
        "purpose",
        "cores",
        "crosssection",
        "footage",
        "completion",
    ]
    plastic = ["ВВГ", "ППГ"]
    design = ["нг", "Пнг"]
    purpose = ["LS", "FRLS", "LTx", "FRLSLTx"]
    cores = range(1, 6)
    crosssection = [0.5, 0.75, 1, 1.5, 2.5, 4, 6, 10]
    footage = {
        0.5: [35000],
        0.75: [30000],
        1: [25000],
        1.5: [20000],
        2.5: [15000],
        4: [10000],
        6: [8000],
        10: [6000],
    }

    file = forms.FileField(label="Загрузка файла")

    @staticmethod
    def validate_date(date):
        if isinstance(date, datetime.datetime):
            return True
        else:
            return False

    @staticmethod
    def get_rows(file):
        workbook = load_workbook(file)
        worksheet = workbook.active
        rows = worksheet.iter_rows(min_row=2, max_col=8, values_only=True)
        return rows

    def clean_file(self):
        file = self.cleaned_data["file"]
        orders_number = Order.objects.all().values_list("batch_number", flat=True)
        if not file.name.endswith(".xlsx"):
            raise forms.ValidationError("Файл не является Excel файлом")
        rows = self.get_rows(file)
        for num, row in enumerate(rows):
            if len(row) != 8:
                raise forms.ValidationError("Должно быть 8 столбцов")
            if not all(row):
                raise forms.ValidationError(
                    f"Должно быть заполнены все ячейки, проверить строку {num+2}"
                )
            if str(row[0]).isdigit() and row[0] in orders_number:
                raise forms.ValidationError(
                    f"Проверте номер заказа {row[0]} в строке {num+2}"
                )
            if (
                row[1] not in self.plastic
                or row[2] not in self.design
                or row[3] not in self.purpose
                or row[4] not in self.cores
                or row[5] not in self.crosssection
            ):
                raise forms.ValidationError(
                    f"Проверь данные заказа {row[1]} {row[2]} {row[3]} {row[4]}x{row[5]} в строке {num+2}"
                )
            if row[6] not in self.footage.get(row[5]):
                raise forms.ValidationError(f"Проверь длину заказа в строке {num+2}")
            if not isinstance(row[7], datetime.datetime):
                raise forms.ValidationError(f"Некорректный формат даты {row[7]} в строке {num+2}")
        return file

    def save(self):
        file = self.cleaned_data["file"]
        rows = self.get_rows(file)
        for row in rows:
            order = Order(
                batch_number=row[0],
                slug=slugify(f"{row[0]}-{row[5]}"),
                plastic=row[1],
                design=row[2],
                purpose=row[3],
                cores=row[4],
                crosssection=row[5],
                footage=row[6],
                completion=row[7],
            )
            order.full_clean()
            order.save()
